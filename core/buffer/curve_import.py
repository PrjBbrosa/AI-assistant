"""CSV / XLSX loader for buffer-block test curves.

openpyxl is imported lazily inside the XLSX reader so importing this module
does not affect application startup for users who only use CSV.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

from core.buffer.calculator import InputError


_DISPLACEMENT_ALIASES = {"x_mm", "displacement_mm", "位移_mm", "x", "displacement"}
_LOADING_ALIASES = {"loading_force_n", "force_loading_n", "加载力_n", "loading", "f_load"}
_UNLOADING_ALIASES = {
    "unloading_force_n",
    "force_unloading_n",
    "卸载力_n",
    "unloading",
    "f_unload",
}
_BRANCH_ALIASES = {"branch", "phase", "曲线"}
_FORCE_ALIASES = {"force_n", "力_n", "force", "f"}

_BRANCH_LOAD_VALUES = {"loading", "load", "加载", "压缩"}
_BRANCH_UNLOAD_VALUES = {"unloading", "unload", "卸载", "回弹"}


def _normalize_header(name: str) -> str:
    return str(name).strip().lstrip("\ufeff").lower()


def _match_column(headers: Sequence[str], aliases: set[str]) -> int:
    lowered = [_normalize_header(header) for header in headers]
    alias_set = {alias.lower() for alias in aliases}
    for index, header in enumerate(lowered):
        if header in alias_set:
            return index
    return -1


def _to_float(value: Any, label: str, row: int) -> float:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError) as exc:
        raise InputError(f"{label} 第 {row} 行不是数字: {value!r}") from exc


def _read_csv_rows(path: Path) -> Tuple[List[str], List[List[str]]]:
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        text = path.read_text(encoding="utf-8", errors="replace")
    if not text.strip():
        raise InputError("CSV 文件为空")
    first_line = text.splitlines()[0] if text.splitlines() else ""
    delimiter = "\t" if "\t" in first_line and "," not in first_line else ","
    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        raise InputError("CSV 文件为空")
    headers = rows[0]
    if not any(str(header).strip() for header in headers):
        raise InputError("CSV 缺少表头")
    data = [row for row in rows[1:] if any(str(cell).strip() for cell in row)]
    return headers, data


def _read_xlsx_rows(path: Path) -> Tuple[List[str], List[List[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise InputError("缺少依赖 openpyxl，请安装 openpyxl 后重试") from exc

    try:
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:
        raise InputError(f"无法打开 XLSX 文件: {exc}") from exc

    try:
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise InputError("XLSX 工作表为空") from exc
        headers = ["" if value is None else str(value) for value in header_row]
        if not any(header.strip() for header in headers):
            raise InputError("XLSX 缺少表头")
        data: List[List[str]] = []
        for row in rows_iter:
            cells = ["" if value is None else str(value) for value in row]
            if any(cell.strip() for cell in cells):
                data.append(cells)
        return headers, data
    finally:
        workbook.close()


def _parse_wide_table(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> Dict[str, Any]:
    x_idx = _match_column(headers, _DISPLACEMENT_ALIASES)
    load_idx = _match_column(headers, _LOADING_ALIASES)
    unload_idx = _match_column(headers, _UNLOADING_ALIASES)
    if x_idx < 0:
        raise InputError("未识别到位移列（支持: x_mm/displacement_mm/位移_mm）")
    if load_idx < 0:
        raise InputError("未识别到加载曲线（支持: loading_force_n/加载力_n 等）")
    if unload_idx < 0:
        raise InputError("未识别到卸载曲线（支持: unloading_force_n/卸载力_n 等）")

    loading: List[Dict[str, float]] = []
    unloading: List[Dict[str, float]] = []
    for row_number, row in enumerate(rows, start=2):
        if len(row) <= max(x_idx, load_idx, unload_idx):
            raise InputError(f"第 {row_number} 行列数不足")
        x = _to_float(row[x_idx], "位移", row_number)
        loading.append({"x_mm": x, "force_n": _to_float(row[load_idx], "加载力", row_number)})
        unloading.append({"x_mm": x, "force_n": _to_float(row[unload_idx], "卸载力", row_number)})

    if not loading:
        raise InputError("未识别到加载曲线")
    if not unloading:
        raise InputError("未识别到卸载曲线")
    return {
        "loading": loading,
        "unloading": unloading,
        "metadata": {
            "format": "wide",
            "rows": len(rows),
            "loading_count": len(loading),
            "unloading_count": len(unloading),
        },
    }


def _parse_long_table(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> Dict[str, Any]:
    branch_idx = _match_column(headers, _BRANCH_ALIASES)
    x_idx = _match_column(headers, _DISPLACEMENT_ALIASES)
    force_idx = _match_column(headers, _FORCE_ALIASES)
    if branch_idx < 0 or x_idx < 0 or force_idx < 0:
        raise InputError("长表必须含 branch / 位移 / 力 三列")

    load_values = {value.lower() for value in _BRANCH_LOAD_VALUES}
    unload_values = {value.lower() for value in _BRANCH_UNLOAD_VALUES}
    loading: List[Dict[str, float]] = []
    unloading: List[Dict[str, float]] = []
    for row_number, row in enumerate(rows, start=2):
        if len(row) <= max(branch_idx, x_idx, force_idx):
            raise InputError(f"第 {row_number} 行列数不足")
        branch = str(row[branch_idx]).strip().lower()
        if branch in load_values:
            target = loading
        elif branch in unload_values:
            target = unloading
        else:
            raise InputError(
                f"第 {row_number} 行 branch 值 {row[branch_idx]!r} 无法识别 "
                "(支持 loading/load/加载/压缩 / unloading/unload/卸载/回弹)"
            )
        target.append(
            {
                "x_mm": _to_float(row[x_idx], "位移", row_number),
                "force_n": _to_float(row[force_idx], "力", row_number),
            }
        )

    if not loading:
        raise InputError("长表未识别到加载曲线")
    if not unloading:
        raise InputError("长表未识别到卸载曲线")
    return {
        "loading": loading,
        "unloading": unloading,
        "metadata": {
            "format": "long",
            "rows": len(rows),
            "loading_count": len(loading),
            "unloading_count": len(unloading),
        },
    }


def _detect_format(headers: Sequence[str]) -> str:
    if (
        _match_column(headers, _LOADING_ALIASES) >= 0
        or _match_column(headers, _UNLOADING_ALIASES) >= 0
    ):
        return "wide"
    if _match_column(headers, _BRANCH_ALIASES) >= 0 or _match_column(headers, _FORCE_ALIASES) >= 0:
        return "long"
    if _match_column(headers, _DISPLACEMENT_ALIASES) < 0:
        raise InputError("未识别到位移列（支持: x_mm/displacement_mm/位移_mm）")
    raise InputError("无法识别表格形态：宽表需加载/卸载力列，长表需 branch + force 列")


def load_buffer_curve(path: Path | str) -> Dict[str, Any]:
    """Load buffer loading/unloading curves from a CSV or XLSX file."""
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        headers, rows = _read_csv_rows(file_path)
    elif suffix == ".xlsx":
        headers, rows = _read_xlsx_rows(file_path)
    else:
        raise InputError(f"文件类型不支持，仅支持 .csv / .xlsx (当前: {suffix})")

    fmt = _detect_format(headers)
    parsed = _parse_wide_table(headers, rows) if fmt == "wide" else _parse_long_table(headers, rows)
    parsed["metadata"]["file_name"] = file_path.name
    return parsed
