"""CSV/XLSX importers for fatigue test data, block spectra, and histories."""

from __future__ import annotations

import csv
import hashlib
from pathlib import Path
from typing import Any, Iterable, Sequence


class ImportError(ValueError):
    """Raised when a fatigue data file cannot be normalized."""


_ALIASES: dict[str, set[str]] = {
    "specimen_id": {"specimen_id", "specimen", "sample_id", "试样编号", "样品编号", "试样号"},
    "level_id": {"level_id", "level", "load_level", "载荷级", "应力级"},
    "condition_group": {"condition_group", "condition", "group", "条件组", "工况组"},
    "stress_amplitude_mpa": {"stress_amplitude_mpa", "stress_amplitude", "amplitude", "sa_mpa", "sa", "应力幅_mpa", "应力幅", "幅值"},
    "stress_mean_mpa": {"stress_mean_mpa", "stress_mean", "mean", "sm_mpa", "sm", "平均应力_mpa", "平均应力", "均值"},
    "stress_max_mpa": {"stress_max_mpa", "stress_max", "maximum", "max", "smax_mpa", "smax", "最大应力_mpa", "最大应力", "最大值"},
    "stress_min_mpa": {"stress_min_mpa", "stress_min", "minimum", "min", "smin_mpa", "smin", "最小应力_mpa", "最小应力", "最小值"},
    "cycles": {"cycles", "cycle_count", "n", "循环数", "加载次数", "循环数n[lw]"},
    "status": {"status", "outcome", "result", "状态", "样品状态"},
    "failure_mode": {"failure_mode", "mode", "失效模式", "失效位置"},
    "time_s": {"time_s", "time", "timestamp", "时间_s", "时间"},
    "value": {"value", "load", "stress", "signal", "载荷", "应力", "信号"},
}

_FAILURE = {"failure", "failed", "fracture", "break", "断裂", "失效", "开裂"}
_RUNOUT = {"runout", "suspended", "survived", "未断裂", "无失效", "通过", "中止"}


def _norm(value: Any) -> str:
    return str(value).strip().lstrip("\ufeff").lower().replace(" ", "_")


def _rows(path: Path, sheet_name: str | None) -> tuple[list[str], list[list[Any]]]:
    if path.suffix.lower() == ".csv":
        text = path.read_text(encoding="utf-8-sig")
        if not text.strip():
            raise ImportError("文件为空")
        first = text.splitlines()[0]
        delimiter = "\t" if "\t" in first and "," not in first else ","
        raw = list(csv.reader(text.splitlines(), delimiter=delimiter))
    elif path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise ImportError("缺少 openpyxl，无法读取 XLSX") from exc
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise ImportError(f"无法打开 XLSX: {exc}") from exc
        try:
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ImportError(f"找不到工作表: {sheet_name}")
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
            raw = [list(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
    else:
        raise ImportError("仅支持 CSV/XLSX")
    raw = [row for row in raw if any(cell is not None and str(cell).strip() for cell in row)]
    if len(raw) < 2:
        raise ImportError("文件缺少表头或数据")
    return ["" if item is None else str(item) for item in raw[0]], raw[1:]


def list_xlsx_sheets(path: Path | str) -> list[str]:
    file_path = Path(path)
    if file_path.suffix.lower() != ".xlsx":
        return []
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _column_map(headers: Sequence[str]) -> dict[str, int]:
    normalized = [_norm(header) for header in headers]
    result: dict[str, int] = {}
    for canonical, aliases in _ALIASES.items():
        choices = {_norm(alias) for alias in aliases}
        for index, header in enumerate(normalized):
            if header in choices:
                result[canonical] = index
                break
    return result


def _cell(row: Sequence[Any], columns: dict[str, int], name: str, default: Any = "") -> Any:
    index = columns.get(name)
    if index is None or index >= len(row) or row[index] is None:
        return default
    return row[index]


def _number(value: Any, label: str, row_number: int) -> float:
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ImportError(f"第 {row_number} 行 {label} 不是数字: {value!r}") from exc


def _status(value: Any, row_number: int) -> str:
    normalized = _norm(value)
    if normalized in {_norm(item) for item in _FAILURE}:
        return "failure"
    if normalized in {_norm(item) for item in _RUNOUT}:
        return "runout"
    raise ImportError(f"第 {row_number} 行状态无法识别: {value!r}")


def _source(path: Path, sheet_name: str | None) -> dict[str, Any]:
    return {
        "file_name": path.name,
        "sheet_name": sheet_name or "(active)",
        "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
    }


def load_sn_test_data(path: Path | str, *, sheet_name: str | None = None) -> dict[str, Any]:
    file_path = Path(path)
    headers, rows = _rows(file_path, sheet_name)
    columns = _column_map(headers)
    if "cycles" not in columns or "status" not in columns:
        raise ImportError("S-N 数据必须包含循环数与状态列")
    has_amp = "stress_amplitude_mpa" in columns
    has_extrema = "stress_max_mpa" in columns and "stress_min_mpa" in columns
    if not has_amp and not has_extrema:
        raise ImportError("S-N 数据必须包含应力幅，或最大/最小应力列")
    specimens: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        cycles = _number(_cell(row, columns, "cycles"), "循环数", row_number)
        if cycles <= 0:
            raise ImportError(f"第 {row_number} 行循环数必须 > 0")
        if has_amp:
            amplitude = _number(_cell(row, columns, "stress_amplitude_mpa"), "应力幅", row_number)
            mean = _number(_cell(row, columns, "stress_mean_mpa", 0), "平均应力", row_number)
        else:
            maximum = _number(_cell(row, columns, "stress_max_mpa"), "最大应力", row_number)
            minimum = _number(_cell(row, columns, "stress_min_mpa"), "最小应力", row_number)
            amplitude = abs(maximum - minimum) / 2.0
            mean = (maximum + minimum) / 2.0
        if amplitude <= 0:
            raise ImportError(f"第 {row_number} 行应力幅必须 > 0")
        specimens.append(
            {
                "specimen_id": str(_cell(row, columns, "specimen_id", row_number)),
                "level_id": str(_cell(row, columns, "level_id", "")),
                "condition_group": str(_cell(row, columns, "condition_group", "")),
                "stress_amplitude_mpa": amplitude,
                "stress_mean_mpa": mean,
                "cycles": cycles,
                "status": _status(_cell(row, columns, "status"), row_number),
                "failure_mode": str(_cell(row, columns, "failure_mode", "")),
            }
        )
    return {"specimens": specimens, "source": _source(file_path, sheet_name)}


def load_spectrum_data(
    path: Path | str, *, kind: str, sheet_name: str | None = None
) -> dict[str, Any]:
    file_path = Path(path)
    headers, rows = _rows(file_path, sheet_name)
    columns = _column_map(headers)
    if kind == "time_series":
        if "value" not in columns:
            raise ImportError("时序必须包含 value/load/stress/载荷/应力列")
        series: list[dict[str, float]] = []
        for row_number, row in enumerate(rows, start=2):
            time_value = _cell(row, columns, "time_s", row_number - 2)
            series.append(
                {
                    "time_s": _number(time_value, "时间", row_number),
                    "value": _number(_cell(row, columns, "value"), "信号值", row_number),
                }
            )
        return {"kind": kind, "series": series, "source": _source(file_path, sheet_name)}
    if kind != "blocks":
        raise ImportError("谱类型必须是 blocks 或 time_series")
    if "cycles" not in columns:
        raise ImportError("块谱必须包含循环数列")
    has_amp = "stress_amplitude_mpa" in columns
    has_extrema = "stress_max_mpa" in columns and "stress_min_mpa" in columns
    if not has_amp and not has_extrema:
        raise ImportError("块谱必须包含幅值/均值或最大/最小值")
    blocks: list[dict[str, float]] = []
    for row_number, row in enumerate(rows, start=2):
        if has_amp:
            amplitude = _number(_cell(row, columns, "stress_amplitude_mpa"), "幅值", row_number)
            mean = _number(_cell(row, columns, "stress_mean_mpa", 0), "均值", row_number)
        else:
            maximum = _number(_cell(row, columns, "stress_max_mpa"), "最大值", row_number)
            minimum = _number(_cell(row, columns, "stress_min_mpa"), "最小值", row_number)
            amplitude = abs(maximum - minimum) / 2.0
            mean = (maximum + minimum) / 2.0
        cycles = _number(_cell(row, columns, "cycles"), "循环数", row_number)
        if amplitude <= 0:
            raise ImportError(f"第 {row_number} 行幅值必须 > 0")
        if cycles <= 0:
            raise ImportError(f"第 {row_number} 行循环数必须 > 0")
        blocks.append(
            {
                "amplitude": amplitude,
                "mean": mean,
                "cycles": cycles,
            }
        )
    return {"kind": kind, "blocks": blocks, "source": _source(file_path, sheet_name)}
